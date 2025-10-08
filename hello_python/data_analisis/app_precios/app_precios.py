import openpyxl as pyxl
from openpyxl import load_workbook 
RUTA =r"C:\Users\Personal\python\Python\myfirstfile.xlsx"

wb = load_workbook(filename= RUTA)

ws = wb.active

#Allegri
allegri_larga = [17.10, 16.60 , 16.50]
allegri_corta = 20.65
allegri_larga_medio = 10.25
allegri_corta_medio = 12.55
allegri_especialidades = 16.30
allegri_pasticho = 18


columna_credito = ws['C4:C9']

#Columna credito 

ws['C4'] = allegri_larga[0]
ws['C5'] = allegri_corta
ws['C6'] = allegri_larga_medio
ws['C7'] = allegri_corta_medio
ws['C8'] = allegri_especialidades
ws['C9'] = allegri_pasticho

#Columna contado

ws['D4'] = allegri_larga
ws['D5'] = allegri_corta
ws['D6'] = allegri_larga_medio
ws['D7'] = allegri_corta_medio
ws['D8'] = allegri_especialidades
ws['D9'] = allegri_pasticho


wb.save('myfirstfile.xlsx')


