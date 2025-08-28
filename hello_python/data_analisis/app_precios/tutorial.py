import openpyxl as pyxl
from openpyxl import Workbook


wb = Workbook()

#Activamos la hoja para poder trabajar en ella
ws = wb.active
#Crea una hoja nueva en el libro
ws1 = wb.create_sheet("my first sheet")

#Para asignar valores a una celda debemos llamarla primero 
a1= ws['A1']

#Luego podremos colocarle el valor que querramos
ws['A1'] = 9*2

wb.save("myfirstfile.xlsx")


